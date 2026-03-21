"""
Forma Pro — SQLite Repository Layer
Logs every extraction to a local database for history, analysis, and reporting.
"""
from __future__ import annotations
import sqlite3
from pathlib import Path
from datetime import datetime
from contextlib import contextmanager

try:
    import openpyxl
except ImportError:
    openpyxl = None

DB_PATH = Path(__file__).parent / "forma_repo.db"

SCHEMA = """
CREATE TABLE IF NOT EXISTS extractions (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    doc_type     TEXT NOT NULL,
    customer     TEXT NOT NULL,
    doc_number   TEXT,
    doc_date     TEXT,
    total_amount REAL DEFAULT 0.0,
    file_path    TEXT,
    output_path  TEXT,
    item_count   INTEGER DEFAULT 0,
    mode         TEXT DEFAULT 'extract',
    created_at   TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS line_items (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    extraction_id  INTEGER NOT NULL REFERENCES extractions(id) ON DELETE CASCADE,
    product_id     TEXT,
    description    TEXT,
    quantity       REAL DEFAULT 0,
    price          REAL DEFAULT 0,
    unit           TEXT,
    amount         REAL DEFAULT 0,
    non_returnable INTEGER DEFAULT 0
);

CREATE INDEX IF NOT EXISTS idx_extractions_customer  ON extractions(customer);
CREATE INDEX IF NOT EXISTS idx_extractions_doc_type  ON extractions(doc_type);
CREATE INDEX IF NOT EXISTS idx_extractions_created   ON extractions(created_at);
CREATE INDEX IF NOT EXISTS idx_line_items_extraction ON line_items(extraction_id);
"""


@contextmanager
def _conn():
    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    try:
        yield con
        con.commit()
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()


def init_db():
    with _conn() as con:
        con.executescript(SCHEMA)


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _to_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(str(val).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _cell(ws, col: int, row: int):
    """Read a cell value, stripping whitespace."""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def _find_header_row(ws, keyword: str, col: int = 1, max_row: int = 20) -> int | None:
    """Find the first row where `col` contains `keyword` (case-insensitive)."""
    kw = keyword.lower()
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and kw in str(v).lower():
            return r
    return None


def _scan_header_labels(ws, label_col: int, value_col: int, max_row: int = 15) -> dict:
    """
    Scan rows 1..max_row looking for key labels in label_col.
    Returns a dict of normalized-label → value.
    """
    result = {}
    for r in range(1, max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        value = ws.cell(row=r, column=value_col).value
        if label and value:
            result[str(label).strip().upper()] = str(value).strip()
    return result


def _parse_quote_xlsx(ws) -> dict:
    """Parse a Quotation Excel workbook."""
    # Header labels are in col D (4), values in col E (5), rows 3-6
    labels = _scan_header_labels(ws, label_col=4, value_col=5, max_row=8)

    customer   = labels.get("CUSTOMER", "")
    doc_number = labels.get("QUOTATION#", labels.get("QUOTE#", ""))
    doc_date   = labels.get("QUOTE DATE", "")

    # Find data start: row after the column-header row that contains "Product ID"
    data_start = None
    for r in range(1, 15):
        v = ws.cell(row=r, column=2).value  # col B
        if v and "product" in str(v).lower():
            data_start = r + 1
            break
    data_start = data_start or 9

    items = []
    total = 0.0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(row):
            continue
        # A=Page(0), B=ProdID(1), C=Desc(2), D=Qty(3), E=Price(4), F=UM(5), G=Amount(6)
        pid  = str(row[1] or "").strip()
        desc = str(row[2] or "").strip()
        if desc.upper() in ("TOTAL", "GRAND TOTAL"):
            total = _to_float(row[6])
            break
        if not pid:
            continue
        items.append({
            "product_id":  pid,
            "description": desc,
            "quantity":    _to_float(row[3]),
            "price":       _to_float(row[4]),
            "unit":        str(row[5] or "").strip(),
            "amount":      _to_float(row[6]),
            "non_returnable": 0,
        })

    if not total and items:
        total = sum(i["amount"] for i in items)

    return {"customer": customer, "doc_number": doc_number, "doc_date": doc_date,
            "total_amount": total, "items": items}


def _parse_po_xlsx(ws) -> dict:
    """Parse a Purchase Order Excel workbook."""
    labels = _scan_header_labels(ws, label_col=4, value_col=5, max_row=12)

    customer   = ""
    doc_number = ""
    doc_date   = ""
    for k, v in labels.items():
        if "CUSTOMER" in k and "PO" not in k and not customer:
            customer = v
        elif "P/O NUMBER" in k or "PO NUMBER" in k or "P\\O NUMBER" in k:
            doc_number = v
        elif "DATE" in k and not doc_date:
            doc_date = v

    # Find header row containing "Product ID"
    data_start = None
    for r in range(1, 20):
        v = ws.cell(row=r, column=2).value
        if v and "product" in str(v).lower():
            data_start = r + 1
            break
    data_start = data_start or 13

    items = []
    total = 0.0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(row):
            continue
        # A=Qty(0), B=PID(1), C=Desc(2), D=SalesOrder(3), E=UM(4), F=Price(5), G=Amount(6), H=ReqDate(7)
        pid  = str(row[1] or "").strip()
        desc = str(row[2] or "").strip()
        if desc.upper() in ("TOTAL", "GRAND TOTAL"):
            total = _to_float(row[6])
            break
        if not pid:
            continue
        items.append({
            "product_id":  pid,
            "description": desc,
            "quantity":    _to_float(row[0]),
            "price":       _to_float(row[5]),
            "unit":        str(row[4] or "").strip(),
            "amount":      _to_float(row[6]),
            "non_returnable": 0,
        })

    if not total and items:
        total = sum(i["amount"] for i in items)

    return {"customer": customer, "doc_number": doc_number, "doc_date": doc_date,
            "total_amount": total, "items": items}


def _parse_invoice_xlsx(ws) -> dict:
    """Parse a Pick-up Ticket / Invoice Excel workbook."""
    labels = _scan_header_labels(ws, label_col=4, value_col=5, max_row=10)

    customer   = ""
    doc_number = ""
    doc_date   = ""
    for k, v in labels.items():
        if "CUSTOMER" in k and not customer:
            customer = v
        elif "ORDER" in k and ("NUMBER" in k or "#" in k) and not doc_number:
            doc_number = v
        elif "DATE" in k and not doc_date:
            doc_date = v

    # Find data start
    data_start = None
    for r in range(1, 15):
        v = ws.cell(row=r, column=2).value
        if v and "product" in str(v).lower():
            data_start = r + 1
            break
    data_start = data_start or 11

    items = []
    total = 0.0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(row):
            continue
        # A=Qty(0), B=PID(1), C=Desc(2), D=UM(3), E=Amount(4)
        pid  = str(row[1] or "").strip()
        desc = str(row[2] or "").strip()
        if desc.upper() in ("TOTAL", "GRAND TOTAL"):
            total = _to_float(row[4])
            break
        if not pid:
            continue
        items.append({
            "product_id":  pid,
            "description": desc,
            "quantity":    _to_float(row[0]),
            "price":       0.0,
            "unit":        str(row[3] or "").strip(),
            "amount":      _to_float(row[4]),
            "non_returnable": 0,
        })

    if not total and items:
        total = sum(i["amount"] for i in items)

    return {"customer": customer, "doc_number": doc_number, "doc_date": doc_date,
            "total_amount": total, "items": items}


def _parse_xlsx(xlsx_path: str, doc_type: str) -> dict:
    if openpyxl is None:
        raise ImportError("openpyxl is required for repo parsing")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    if doc_type == "quote":
        return _parse_quote_xlsx(ws)
    elif doc_type == "po":
        return _parse_po_xlsx(ws)
    elif doc_type == "invoice":
        return _parse_invoice_xlsx(ws)
    else:
        return {"customer": "UNKNOWN", "doc_number": "", "doc_date": "",
                "total_amount": 0.0, "items": []}


# ---------------------------------------------------------------------------
# Write operations
# ---------------------------------------------------------------------------

def log_from_xlsx(xlsx_path: str, doc_type: str, pdf_path: str = "",
                  mode: str = "extract") -> int:
    """Parse an Excel output and insert it into the repo. Returns extraction_id."""
    init_db()
    data = _parse_xlsx(xlsx_path, doc_type)
    now  = datetime.now().isoformat(timespec="seconds")

    with _conn() as con:
        cur = con.execute(
            """INSERT INTO extractions
               (doc_type, customer, doc_number, doc_date, total_amount,
                file_path, output_path, item_count, mode, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (doc_type,
             data["customer"] or "UNKNOWN",
             data["doc_number"],
             data["doc_date"],
             data["total_amount"],
             pdf_path,
             xlsx_path,
             len(data["items"]),
             mode,
             now),
        )
        eid = cur.lastrowid
        for item in data["items"]:
            con.execute(
                """INSERT INTO line_items
                   (extraction_id, product_id, description, quantity, price, unit, amount, non_returnable)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (eid, item["product_id"], item["description"],
                 item["quantity"], item["price"], item["unit"],
                 item["amount"], item.get("non_returnable", 0)),
            )
    return eid


def delete_extraction(extraction_id: int) -> bool:
    init_db()
    with _conn() as con:
        con.execute("DELETE FROM extractions WHERE id = ?", (extraction_id,))
    return True


# ---------------------------------------------------------------------------
# Read operations
# ---------------------------------------------------------------------------

def get_stats() -> dict:
    init_db()
    with _conn() as con:
        total = con.execute("SELECT COUNT(*) FROM extractions").fetchone()[0]
        by_type = {
            r[0]: r[1]
            for r in con.execute(
                "SELECT doc_type, COUNT(*) FROM extractions GROUP BY doc_type"
            ).fetchall()
        }
        total_value = con.execute(
            "SELECT COALESCE(SUM(total_amount), 0) FROM extractions"
        ).fetchone()[0]
        customers = con.execute(
            "SELECT COUNT(DISTINCT customer) FROM extractions"
        ).fetchone()[0]
        this_month = con.execute(
            "SELECT COUNT(*) FROM extractions "
            "WHERE strftime('%Y-%m', created_at) = strftime('%Y-%m', 'now')"
        ).fetchone()[0]
        top_customers = con.execute(
            """SELECT customer, COUNT(*) as cnt, SUM(total_amount) as val
               FROM extractions GROUP BY customer ORDER BY val DESC LIMIT 5"""
        ).fetchall()

    return {
        "total_extractions": total,
        "by_type":           by_type,
        "total_value":       round(total_value, 2),
        "customers":         customers,
        "this_month":        this_month,
        "top_customers":     [dict(r) for r in top_customers],
    }


def get_recent(limit: int = 15) -> list:
    init_db()
    with _conn() as con:
        rows = con.execute(
            """SELECT id, doc_type, customer, doc_number, doc_date,
                      total_amount, output_path, created_at, item_count
               FROM extractions ORDER BY created_at DESC LIMIT ?""",
            (limit,),
        ).fetchall()
    return [dict(r) for r in rows]


def get_all_extractions(doc_type: str = None, customer: str = None,
                         limit: int = 50, offset: int = 0) -> dict:
    init_db()
    where, params = [], []
    if doc_type:
        where.append("doc_type = ?");  params.append(doc_type)
    if customer:
        where.append("customer LIKE ?"); params.append(f"%{customer}%")
    w = ("WHERE " + " AND ".join(where)) if where else ""

    with _conn() as con:
        total = con.execute(f"SELECT COUNT(*) FROM extractions {w}", params).fetchone()[0]
        rows  = con.execute(
            f"""SELECT id, doc_type, customer, doc_number, doc_date,
                       total_amount, output_path, created_at, item_count
                FROM extractions {w} ORDER BY created_at DESC LIMIT ? OFFSET ?""",
            params + [limit, offset],
        ).fetchall()
    return {"total": total, "rows": [dict(r) for r in rows]}


def get_extraction(extraction_id: int) -> dict | None:
    init_db()
    with _conn() as con:
        row = con.execute(
            "SELECT * FROM extractions WHERE id = ?", (extraction_id,)
        ).fetchone()
        if not row:
            return None
        items = con.execute(
            "SELECT * FROM line_items WHERE extraction_id = ? ORDER BY id",
            (extraction_id,),
        ).fetchall()
    result = dict(row)
    result["items"] = [dict(i) for i in items]
    return result


def get_customers() -> list:
    init_db()
    with _conn() as con:
        rows = con.execute(
            """SELECT customer,
                      COUNT(*) as doc_count,
                      COALESCE(SUM(total_amount), 0) as total_value,
                      MAX(created_at) as last_activity,
                      SUM(CASE WHEN doc_type='quote'   THEN 1 ELSE 0 END) as quotes,
                      SUM(CASE WHEN doc_type='po'      THEN 1 ELSE 0 END) as pos,
                      SUM(CASE WHEN doc_type='invoice' THEN 1 ELSE 0 END) as invoices
               FROM extractions
               GROUP BY customer
               ORDER BY last_activity DESC"""
        ).fetchall()
    return [dict(r) for r in rows]


def get_customer_extractions(customer: str) -> list:
    init_db()
    with _conn() as con:
        rows = con.execute(
            """SELECT id, doc_type, doc_number, doc_date, total_amount,
                      output_path, created_at, item_count
               FROM extractions WHERE customer = ? ORDER BY created_at DESC""",
            (customer,),
        ).fetchall()
    return [dict(r) for r in rows]


def get_top_products(limit: int = 20) -> list:
    init_db()
    with _conn() as con:
        rows = con.execute(
            """SELECT product_id, description,
                      COUNT(*) as appearances,
                      SUM(quantity) as total_qty,
                      SUM(amount) as total_value
               FROM line_items
               WHERE product_id != '' AND product_id IS NOT NULL
               GROUP BY product_id
               ORDER BY total_value DESC LIMIT ?""",
            (limit,),
        ).fetchall()
    return [dict(r) for r in rows]
