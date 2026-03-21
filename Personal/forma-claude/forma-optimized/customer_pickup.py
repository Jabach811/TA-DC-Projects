"""
customer_pickup.py  —  OPTIMISED VERSION
Original: Forma - Main Folder/customer_pickup.py

What changed vs the original
──────────────────────────────
1. PARALLEL OCR  — ThreadPoolExecutor runs Tesseract on all pages concurrently.
   On a multi-page pick ticket this cuts OCR time from ~N*4 s to ~4 s.
2. DEAD CODE REMOVED — _search_pattern() had an unreachable second copy of its
   loop body (lines 232-236 in the original, after a return statement).
Core extraction logic is IDENTICAL to the original.
"""

from __future__ import annotations

import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Sequence

import pandas as pd
import pytesseract
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pdf2image import convert_from_path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import LOGO_IMAGE, POPPLER_PATH, TESSERACT_PATH, USER_DATA_ROOT, ensure_dir

__all__ = ["customer_pickup"]

# ── Workbook layout helpers ──────────────────────────────────────────────────

HEADER_LABELS = [
    ("Customer",     "customer"),
    ("Ship Via",     "ship_via"),
    ("Order Number", "order_number"),
    ("Customer PO",  "customer_po"),
    ("Order Date",   "order_date"),
    ("Contact",      "contact"),
    ("Territory",    "territory"),
]

# ── Module-level regex (original already had most of these here) ─────────────

ORDER_NUMBER_PATTERN = re.compile(r"\b\d{3}[A-Za-z]{2}\d{3}-\d+\b")
PRO_SINGLE     = re.compile(r'\bPRO\s*[-"]?\s*90\b', re.IGNORECASE)
PRO_PAIR       = re.compile(r'\bPRO([0-9lLiI|]+)x(PRO\d+)', re.IGNORECASE)
PRO_DOUBLE     = re.compile(r'\b(PRO\d{1,4})[^0-9]{1,5}(PRO\d{1,4})\b', re.IGNORECASE)
PRO_L_MIX      = re.compile(r'\bPRO[1lI|]{1,2}50\b', re.IGNORECASE)
PRO_90_VARIANT = re.compile(r'\bPRO[-\s]*9[0O](?:[0O9])?\b', re.IGNORECASE)
PRO_NUM_FIX    = re.compile(r'[lLiI|]')
PRO_RATIO      = re.compile(r'(["\d])\s*[%xX*]+\s*(\d+/?\d*")')
PRO_MULTI_X    = re.compile(r'([xX])[xX]+')
PRO_X_BETWEEN  = re.compile(r'(\d)x[Xx*%]+(?=\d)')
PRO_SPACE_DASH = re.compile(r'\bx\s*-\s*([A-Za-z0-9]+)')

ANSI1          = re.compile(r'ANTSI', re.IGNORECASE)
ANSI2          = re.compile(r'ANST[1iIl|]', re.IGNORECASE)
PSI_MERGE      = re.compile(r'(\d)PST[1iIl|]', re.IGNORECASE)
PSI_WORD       = re.compile(r'\bPST[1iIl|]\b', re.IGNORECASE)
PSI_FIX        = re.compile(r'(\d)PSI\b', re.IGNORECASE)
SDR_PACK       = re.compile(r'\bSDR[1ilI|]{1,6}\b', re.IGNORECASE)
SDR_SPACE_PACK = re.compile(r'\bSDR\s+[1ilI|]{1,6}\b', re.IGNORECASE)
SCH80_BAD      = re.compile(r'\bSCH80[^0-9\s]{1,3}\b', re.IGNORECASE)
SCH40_BAD      = re.compile(r'\bSCH40[^0-9\s]{1,3}\b', re.IGNORECASE)
PN_BAD         = re.compile(r'\(\s*P\s*N\s*[6GHB9QOD]\s*\)', re.IGNORECASE)
PN_JOIN        = re.compile(r'(?<!\s)(PN6)', re.IGNORECASE)
PN_TRAIL9      = re.compile(r'\bPN([6GHB9QOD])9\b', re.IGNORECASE)
PN_LEADING9    = re.compile(r'\b9\s*(PN6)\b', re.IGNORECASE)

PRO_BAD_VARIANTS = (
    "PROQO","PROQ0","PRO-90","PRO90","PROQ90","PROYQO",
    "PROY9O","PROQ9O","PRO-QO","PRO-Q0","PRO9O","PROO9O",
    "PROYO","PRO-YO","PROY0","PRO-Y0",
)
PRO_BAD_PATTERN = re.compile(
    "|".join(re.escape(v) for v in PRO_BAD_VARIANTS), re.IGNORECASE
)

PRO_SLIP = re.compile(r'\bSLIP\b', re.IGNORECASE)
PID_SUFFIX = re.compile(r'([A-Za-z0-9])$')

NON_RETURNABLE_PHRASE = "PRODUCT IS NON-RETURNABLE ITEMS MAY NOT BE CANCELLED"
NON_RETURNABLE_LINES  = {"PRODUCT IS NON-RETURNABLE", "ITEMS MAY NOT BE CANCELLED"}


# ── Data container ───────────────────────────────────────────────────────────

@dataclass(slots=True)
class ProductRow:
    quantity:   int | str
    product_id: str
    description: str
    unit:       str
    amount:     float

    def as_record(self) -> dict[str, object]:
        return {
            "Quantity":   self.quantity,
            "Product ID": self.product_id,
            "Description": self.description,
            "U/M":        self.unit,
            "Amount":     self.amount,
        }


# ── Public API ───────────────────────────────────────────────────────────────

def customer_pickup(file_path: str, output_root: Path | None = None) -> Path:
    """Convert a Harrington customer pick-up ticket PDF into a styled Excel workbook."""
    pdf_file = Path(file_path).resolve()
    if not pdf_file.exists():
        raise FileNotFoundError(pdf_file)

    pytesseract.pytesseract.tesseract_cmd = str(TESSERACT_PATH)
    ocr_lines = _ocr_pdf(pdf_file)

    header                          = _parse_header(ocr_lines)
    raw_products, non_returnable_flags = _parse_products(ocr_lines)
    products                        = _clean_products(raw_products)

    total_amount = sum(row.amount for row in products)
    products.append(ProductRow("", "", "TOTAL", "", total_amount))
    non_returnable_flags.append(False)

    df = pd.DataFrame([row.as_record() for row in products])

    output_root   = ensure_dir(Path(output_root)) if output_root else ensure_dir(USER_DATA_ROOT)
    customer_name = _sanitize_filename(header.get("customer", "UNKNOWN"))
    order_number  = _sanitize_filename(header.get("order_number", "UNKNOWN"))
    ship_via      = _sanitize_filename(_normalize_ship_via(header.get("ship_via", "") or "UNKNOWN"))

    destination    = ensure_dir(Path(output_root) / customer_name / "Pick Ticket Orders")
    workbook_path  = destination / f"{order_number}{('_' + ship_via) if ship_via else ''}.xlsx"

    df.to_excel(workbook_path, index=False, startrow=9)
    _format_workbook(workbook_path, header, non_returnable_flags)
    return workbook_path


# ── OCR — parallel ───────────────────────────────────────────────────────────

def _ocr_worker(img) -> str:
    """OCR one page in a thread (Tesseract subprocess releases the GIL)."""
    return pytesseract.image_to_string(img, config="--oem 3 --psm 6")


def _ocr_pdf(pdf_file: Path) -> List[str]:
    """Render PDF and OCR all pages concurrently."""
    images      = convert_from_path(pdf_file, dpi=300, poppler_path=str(POPPLER_PATH))
    max_workers = min(len(images), os.cpu_count() or 4, 4)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        page_texts = list(executor.map(_ocr_worker, images))
        # executor.map preserves order — no sort needed

    lines: List[str] = []
    for text in page_texts:
        lines.extend(seg.strip() for seg in text.splitlines() if seg.strip())
    return lines


# ── Header parsing ───────────────────────────────────────────────────────────

def _parse_header(lines: Sequence[str]) -> dict[str, str]:
    header = {
        "customer":    "UNKNOWN",
        "ship_via":    "",
        "order_number": "UNKNOWN",
        "customer_po": "UNKNOWN",
        "order_date":  "",
        "contact":     "",
        "territory":   "",
    }

    for idx, line in enumerate(lines):
        if re.fullmatch(r"\d{6}", line):
            if idx + 1 < len(lines):
                header["customer"] = lines[idx + 1]
            break

    order_number = _search_pattern(lines, ORDER_NUMBER_PATTERN)
    if not order_number:
        order_number = _search_first(lines, r"Order#\s*:\s*([^\s]+)") or header["order_number"]
    order_number           = order_number.replace('O', '0').replace('o', '0')
    header["order_number"] = order_number

    customer_po = _search_first(lines, r"CustP\/[0O]\s*:\s*([^\s]+)")
    if customer_po:
        header["customer_po"] = customer_po

    header["order_date"] = _search_first(lines, r"Ord-Date\s*:?\s*([0-9/]+)") or ""

    contact = _search_first(lines, r"Contact:?\s*([A-Z][A-Z\s]+?)\s+Written\b")
    if contact:
        header["contact"] = contact.title()

    header["territory"] = _search_first(lines, r"Terr:?\s*([A-Z0-9]+)") or ""

    ship_via = _extract_ship_via(lines)
    if ship_via:
        header["ship_via"] = ship_via

    return header


def _search_first(lines: Sequence[str], pattern: str) -> str | None:
    regex = re.compile(pattern, re.IGNORECASE)
    for line in lines:
        m = regex.search(line)
        if m:
            return m.group(1).strip()
    return None


def _search_pattern(lines: Sequence[str], pattern: re.Pattern[str]) -> str:
    """Return the first match of a compiled pattern across all lines."""
    for line in lines:
        m = pattern.search(line)
        if m:
            return m.group(0).upper()
    return ""
    # NOTE: the original had an unreachable duplicate of the loop body here
    # (lines 232-236).  Removed.


def _normalize_ship_via(ship_via: str) -> str:
    cleaned = (ship_via or "").strip()
    if cleaned.upper().startswith("COMPANY"):
        return "Company Truck"
    return cleaned


def _sanitize_filename(raw: str) -> str:
    cleaned = re.sub(r"[^\w\- ]", "_", (raw or "UNKNOWN").strip())
    return cleaned or "UNKNOWN"


def _extract_ship_via(lines: Sequence[str]) -> str:
    for line in lines:
        lower = line.lower()
        if "ship" not in lower or "via" not in lower:
            continue
        before_fob = re.split(r"\bFOB\b", line, flags=re.IGNORECASE)[0]
        m = re.search(r"Ship\s*[- ]?\s*Via\s*:?\s*(.*)", before_fob, re.IGNORECASE)
        if not m:
            continue
        text = re.sub(r"\s+", " ", m.group(1).strip().lower())
        if (("customer" in text and any(f in text for f in ("pick up", "pickup", "pick-up")))
                or "will call" in text or "willcall" in text):
            return "Customer Pick Up"
        if "company" in text and "truck" in text:
            return "Company Truck"
        if "common" in text and "carrier" in text:
            return "Common Carrier"
        if "ups" in text:
            return "UPS"
    return ""


# ── Product parsing ──────────────────────────────────────────────────────────

def _extract_inline_description(remainder: str) -> str:
    tokens: list[str] = []
    for token in remainder.split():
        normalized = token.replace(",", "")
        if re.fullmatch(r"[\d\.()]+", normalized):
            break
        if normalized.upper() in {"EA", "FT", "BA", "EBA", "FA", "EFA"}:
            break
        tokens.append(token)
    return " ".join(tokens)


def _strip_non_returnable(lines: List[str]) -> tuple[List[str], bool]:
    flag     = False
    combined = " ".join(e.upper().strip() for e in lines)
    if NON_RETURNABLE_PHRASE in combined:
        flag = True
    cleaned: List[str] = []
    for entry in lines:
        upper = entry.upper().strip()
        if upper in NON_RETURNABLE_LINES or (upper and NON_RETURNABLE_PHRASE in upper):
            flag = True
            continue
        cleaned.append(entry)
    if flag:
        cleaned = [e for e in cleaned if e.strip()]
    return cleaned, flag


def _parse_products(lines: Sequence[str]) -> tuple[List[ProductRow], List[bool]]:
    products: List[ProductRow] = []
    flags:    List[bool]       = []

    product_line_regex = re.compile(
        r"^(?P<pid>[A-Z0-9][A-Z0-9\-\.]*\d[A-Z0-9\-\.]*)\s+(?P<qty>\d+)\s+(?P<um>[A-Z/]+)"
    )

    i = 0
    while i < len(lines):
        line  = lines[i]
        match = product_line_regex.match(line)
        if not match:
            i += 1
            continue

        pid  = match.group("pid")
        qty  = int(match.group("qty"))
        unit = match.group("um").strip().upper()
        if unit in {"BA", "EBA", "FA", "EFA"}:
            unit = "EA"

        remainder        = line[match.end():].strip()
        description_lines: List[str] = []
        inline_desc = _extract_inline_description(remainder) if remainder else ""
        if inline_desc:
            description_lines.append(inline_desc)

        amount_text = ""
        i += 1
        while i < len(lines):
            next_line = lines[i]
            if product_line_regex.match(next_line) or next_line.startswith("In Frt"):
                break
            if re.search(r"BIN:", next_line, re.IGNORECASE):
                for token in reversed(next_line.split()):
                    candidate = token.replace(",", "")
                    if re.fullmatch(r"\d+(?:\.\d+)?", candidate):
                        amount_text = candidate
                        break
                i += 1
                break
            description_lines.append(next_line)
            i += 1

        description_lines, non_returnable = _strip_non_returnable(description_lines)
        description = " ".join(description_lines).strip()
        amount      = float(amount_text) if amount_text else 0.0

        products.append(ProductRow(qty, pid, description, unit, amount))
        flags.append(non_returnable)

    return products, flags


def _clean_products(rows: Iterable[ProductRow]) -> List[ProductRow]:
    cleaned: List[ProductRow] = []
    for row in rows:
        description = row.description.strip()
        product_id  = row.product_id.strip()
        if not description and not product_id:
            cleaned.append(row)
            continue

        product_id = re.sub(r"(?<=\d)\s+(?=\d)", "", product_id)
        desc_upper = description.upper()

        if "ASAHT" in desc_upper:
            description = description.replace("ASAHT", "ASAHI")
            desc_upper  = description.upper()

        if "ANTSI" in desc_upper:
            description = ANSI1.sub("ANSI", description)
            desc_upper  = description.upper()
        if "ANST" in desc_upper:
            description = ANSI2.sub("ANSI", description)
            desc_upper  = description.upper()

        if "PST" in desc_upper:
            description = PSI_MERGE.sub(r"\1 PSI", description)
            description = PSI_WORD.sub("PSI", description)
            description = PSI_FIX.sub(r"\1 PSI", description)
            desc_upper  = description.upper()

        if "SDR" in desc_upper:
            description = SDR_PACK.sub("SDR11", description)
            description = SDR_SPACE_PACK.sub("SDR11", description)
            desc_upper  = description.upper()

        if "SCH80" in desc_upper:
            description = SCH80_BAD.sub("SCH80", description)
            desc_upper  = description.upper()
        if "SCH40" in desc_upper:
            description = SCH40_BAD.sub("SCH40", description)
            desc_upper  = description.upper()

        if "PN" in desc_upper:
            description = PN_BAD.sub(" PN6", description)
            description = PN_JOIN.sub(r" \1", description)
            description = PN_TRAIL9.sub("PN6", description)
            description = PN_LEADING9.sub(r"\1", description)
            desc_upper  = description.upper()

        if "PRO" in desc_upper:
            description = PRO_BAD_PATTERN.sub("PRO-90", description)
            description = PRO_SINGLE.sub("PRO-90", description)
            description = PRO_PAIR.sub(
                lambda m: f"PRO{PRO_NUM_FIX.sub('1', m.group(1))}x{m.group(2)}", description
            )
            description = PRO_DOUBLE.sub(r"\1x\2", description)
            description = PRO_90_VARIANT.sub("PRO-90", description)
            desc_upper  = description.upper()

        description = PRO_RATIO.sub(r"\1x\2", description)
        description = PRO_MULTI_X.sub(r"\1", description)
        description = PRO_X_BETWEEN.sub(r"\1x", description)
        description = PRO_SPACE_DASH.sub(r"x\1", description)
        description = description.replace("(", "").replace(")", "")
        description = re.sub(r"\s+", " ", description).strip()

        if PRO_L_MIX.search(product_id):
            product_id = PRO_L_MIX.sub("PRO150", product_id)
        if PRO_90_VARIANT.search(product_id):
            product_id = PRO_90_VARIANT.sub("PRO90", product_id)

        if "FLANGE BLIND" in desc_upper and product_id.endswith("8"):
            product_id = f"{product_id[:-1]}S"
        if (" S " in desc_upper or PRO_SLIP.search(desc_upper)) and not product_id.endswith("S"):
            product_id = PID_SUFFIX.sub("S", product_id)

        cleaned.append(ProductRow(
            quantity=row.quantity,
            product_id=product_id,
            description=description,
            unit=row.unit,
            amount=row.amount,
        ))

    return cleaned


# ── Workbook styling (unchanged) ─────────────────────────────────────────────

def _format_workbook(
    path: Path, header: dict[str, str], non_returnable_flags: Sequence[bool]
) -> None:
    wb = load_workbook(path)
    ws = wb.active

    for column, width in {"A": 25, "B": 25, "C": 106, "D": 25, "E": 25}.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[9].height = 30

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for cell in ws[10]:
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = header_fill
        cell.border    = border

    currency_format = '"$"#,##0.00'
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, max_col=5):
        for idx, cell in enumerate(row):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
            if idx == 4:
                cell.number_format = currency_format

    non_returnable_fill = PatternFill(start_color="FFECC6", end_color="FFECC6", fill_type="solid")
    for row_idx, row in enumerate(ws.iter_rows(min_row=11, max_row=ws.max_row, max_col=5)):
        if non_returnable_flags and row_idx < len(non_returnable_flags) and non_returnable_flags[row_idx]:
            for cell in row:
                cell.fill = non_returnable_fill

    total_row_index = ws.max_row
    total_label_cell = ws.cell(row=total_row_index, column=3)
    if (total_label_cell.value or "").strip().upper() == "TOTAL":
        for col_idx in range(1, 6):
            cell = ws.cell(row=total_row_index, column=col_idx)
            cell.font = Font(bold=True)
            if col_idx == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A11"
    ws.sheet_view.showGridLines = False

    for offset, (label, key) in enumerate(HEADER_LABELS):
        row_idx = 2 + offset
        ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=7)
        label_cell            = ws[f"D{row_idx}"]
        label_cell.value      = label
        label_cell.font       = Font(bold=True)
        label_cell.alignment  = Alignment(horizontal="right", vertical="center")
        value_cell            = ws[f"E{row_idx}"]
        value_cell.value      = header.get(key, "")
        value_cell.alignment  = Alignment(horizontal="left", vertical="center")

    if LOGO_IMAGE and LOGO_IMAGE.exists():
        logo        = XLImage(str(LOGO_IMAGE))
        logo.width  *= 0.82
        logo.height *= 0.82
        logo.anchor  = "C1"
        ws.add_image(logo)

    wb.save(path)


# ── CLI entry point ──────────────────────────────────────────────────────────

if __name__ == "__main__":  # pragma: no cover
    import argparse
    parser = argparse.ArgumentParser(description="Customer pick-up PDF extractor")
    parser.add_argument("pdf",     help="Path to the customer pick-up PDF")
    parser.add_argument("--output", help="Optional output directory", default=None)
    args   = parser.parse_args()
    output = customer_pickup(args.pdf, Path(args.output) if args.output else None)
    print(f"Customer pick-up exported to {output}")
