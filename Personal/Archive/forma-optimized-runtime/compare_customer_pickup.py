from __future__ import annotations

import importlib.util
import sys
import time
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
FORMA_ROOT = Path(r"C:\Users\mabac\OneDrive\Desktop\Joel\Hipco Project\Forma -  Main Folder")
DEFAULT_SAMPLE_PDF = FORMA_ROOT / "20250924132515308.pdf"
OUTPUT_ROOT = ROOT / "compare-pickup-output"


def load_module(module_name: str, path: Path):
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def read_rows(path: Path) -> list[list[object]]:
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def normalize_rows(path: Path) -> list[list[object]]:
    rows = read_rows(path)
    trimmed = []
    for row in rows[9:]:
        materialized = list(row[:5])
        if not any(value is not None and value != "" for value in materialized):
            continue
        trimmed.append(materialized)
    return trimmed


def compare_workbooks(left: Path, right: Path) -> tuple[bool, str]:
    left_values = normalize_rows(left)
    right_values = normalize_rows(right)
    if left_values == right_values:
        return True, "Workbook pickup rows match exactly."

    max_rows = max(len(left_values), len(right_values))
    for row_idx in range(max_rows):
        left_row = left_values[row_idx] if row_idx < len(left_values) else []
        right_row = right_values[row_idx] if row_idx < len(right_values) else []
        if left_row != right_row:
            return False, f"First mismatch at row {row_idx + 11}: original={left_row} optimized={right_row}"
    return False, "Workbook comparison failed for an unknown reason."


def main() -> None:
    sys.path.insert(0, str(FORMA_ROOT))
    sample_pdf = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SAMPLE_PDF
    original = load_module("forma_original_customer_pickup", FORMA_ROOT / "customer_pickup.py")
    optimized = load_module("forma_optimized_customer_pickup", ROOT / "customer_pickup.py")

    original_output_root = OUTPUT_ROOT / "original"
    optimized_output_root = OUTPUT_ROOT / "optimized"
    original_output_root.mkdir(parents=True, exist_ok=True)
    optimized_output_root.mkdir(parents=True, exist_ok=True)

    start = time.perf_counter()
    original_path = Path(original.customer_pickup(str(sample_pdf), output_root=str(original_output_root)))
    original_seconds = time.perf_counter() - start

    start = time.perf_counter()
    optimized_path = Path(optimized.customer_pickup(str(sample_pdf), output_root=str(optimized_output_root)))
    optimized_seconds = time.perf_counter() - start

    matches, message = compare_workbooks(original_path, optimized_path)

    print(f"Original output:  {original_path}")
    print(f"Optimized output: {optimized_path}")
    print(f"Original time:    {original_seconds:.2f}s")
    print(f"Optimized time:   {optimized_seconds:.2f}s")
    print(f"Comparison:       {message}")
    if original_seconds:
        improvement = ((original_seconds - optimized_seconds) / original_seconds) * 100
        print(f"Runtime delta:    {improvement:.1f}%")
    print(f"Matched values:   {matches}")


if __name__ == "__main__":
    main()
