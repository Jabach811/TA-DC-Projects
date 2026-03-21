from pathlib import Path
import re
QUOTE_NUMBER_PATTERN = re.compile(r"\b\d{3}[A-Za-z]\d{4}\b")

def _normalize_quote_number(value: str) -> str:
    if not value:
        return ""
    match = QUOTE_NUMBER_PATTERN.search(value)
    if match:
        return match.group(0).upper()
    cleaned = re.sub(r"[^A-Za-z0-9]", "", value).upper()
    fallback = re.search(r"\d{3}[A-Za-z]\d{4}", cleaned)
    if fallback:
        return fallback.group(0)
    zero_variant = re.search(r"\d{3}0\d{4}", cleaned)
    if zero_variant:
        candidate = zero_variant.group(0)
        return candidate[:3] + 'O' + candidate[4:]
    return value.upper()
from config import (
    TESSERACT_PATH,
    POPPLER_PATH as CFG_POPPLER_PATH,
    LOGO_IMAGE as CFG_LOGO_IMAGE,
    USER_DATA_ROOT,
    QUOTES_DIR,
    ensure_dir
)
from ocr_service import OCRBundle, get_ocr_bundle

def quotation_extract(file_path, output_root=None, ocr_bundle: OCRBundle | None = None):

    import re
    import pandas as pd
    from pdf2image import convert_from_path
    import pytesseract
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    # --- Config ---
    PDF_FILE = Path(file_path)
    POPPLER_PATH = str(CFG_POPPLER_PATH)
    LOGO_PATH = CFG_LOGO_IMAGE
    OUTPUT_ROOT = ensure_dir(Path(output_root)) if output_root else ensure_dir(USER_DATA_ROOT)
    pytesseract.pytesseract.tesseract_cmd = str(TESSERACT_PATH)
    all_products = []
    # --- Regex helpers (compile once) ---
    PRO_SINGLE = re.compile(r'\bPRO\s*[-"]?\s*90\b', re.IGNORECASE)
    PRO_PAIR = re.compile(r'\bPRO([0-9lLiI|]+)x(PRO\d+)', re.IGNORECASE)
    PRO_DOUBLE = re.compile(r'\b(PRO\d{1,4})[^0-9]{1,5}(PRO\d{1,4})\b', re.IGNORECASE)
    PRO_L_MIX = re.compile(r'\bPRO[1lI|]{1,2}50\b', re.IGNORECASE)
    PRO_90_VARIANT = re.compile(r'\bPRO[-\s]*9[0O](?:[0O9])?\b', re.IGNORECASE)
    PRO_NUM_FIX = re.compile(r'[lLiI|]')

    ANSI1 = re.compile(r'ANTSI', re.IGNORECASE)
    ANSI2 = re.compile(r'ANST[1iIl|]', re.IGNORECASE)
    PSI_MERGE = re.compile(r'(\d)PST[1iIl|]', re.IGNORECASE)
    PSI_WORD = re.compile(r'\bPST[1iIl|]\b', re.IGNORECASE)
    PSI_FIX = re.compile(r'(\d)PSI\b', re.IGNORECASE)

    SDR_PACK = re.compile(r'\bSDR[1ilI|]{1,6}\b', re.IGNORECASE)
    SDR_SPACE_PACK = re.compile(r'\bSDR\s+[1ilI|]{1,6}\b', re.IGNORECASE)

    SCH80_BAD = re.compile(r'\bSCH80[^0-9\s]{1,3}\b', re.IGNORECASE)
    SCH40_BAD = re.compile(r'\bSCH40[^0-9\s]{1,3}\b', re.IGNORECASE)

    PN_BAD = re.compile(r'\(\s*P\s*N\s*[6GHB9QOD]\s*\)', re.IGNORECASE)
    PN_JOIN = re.compile(r'(?<!\s)(PN6)', re.IGNORECASE)
    PN_TRAIL9 = re.compile(r'\bPN([6GHB9QOD])9\b', re.IGNORECASE)
    PN_LEADING9 = re.compile(r'\b9\s*(PN6)\b', re.IGNORECASE)

    PRO_BAD_VARIANTS = (
        'PROQO', 'PROQ0', 'PRO-90', 'PRO90', 'PROQ90', 'PROYQO',
        'PROY9O', 'PROQ9O', 'PRO-QO', 'PRO-Q0', 'PRO9O', 'PROO9O',
        'PROYO', 'PRO-YO', 'PROY0', 'PRO-Y0'
    )
    PRO_BAD_PATTERN = re.compile('|'.join(re.escape(v) for v in PRO_BAD_VARIANTS), re.IGNORECASE)

    PRO_SLIP = re.compile(r'\bSLIP\b', re.IGNORECASE)

    NON_RETURNABLE_PHRASE = "PRODUCT IS NON-RETURNABLE ITEMS MAY NOT BE CANCELLED"
    NON_RETURNABLE_LINES = {"PRODUCT IS NON-RETURNABLE", "ITEMS MAY NOT BE CANCELLED"}

    DIM_RATIO = re.compile(r'([\"\d])\s*[%xX*]+\s*(\d+/?\d*\")')
    DIM_MULTI_X = re.compile(r'([xX])[xX]+')
    DIM_X_BETWEEN = re.compile(r'(\d)x[Xx*%]+(?=\d)')
    DIM_SPACE_DASH = re.compile(r'\bx\s*-\s*([A-Za-z0-9]+)')

    PID_SUFFIX = re.compile(r'([A-Za-z0-9])$')

    OCR_JUNK = re.compile(r'^(?:[KRE\s]{3,})+$', re.IGNORECASE)

    JUNK_SUFFIX = re.compile(r'(?: - )?(?:[A-Z]{2,5}\s+){4,}[A-Z]{2,5}$', re.IGNORECASE)
    JUNK_SUFFIX_STRIPPER = re.compile(r'(?:\s*-\s*(?:[A-Z]{2,6})){4,}\s*$', re.IGNORECASE)
    CONTINUE_PAT = re.compile(r"(?i)\bcontinue(?:d)?\b")



    # --- Helper Functions ---

    def strip_non_returnable(lines):
        flag = False
        filtered = []
        combined = " ".join(entry.upper().strip() for entry in lines)
        if NON_RETURNABLE_PHRASE in combined:
            flag = True
        for entry in lines:
            upper = entry.upper().strip()
            if upper in NON_RETURNABLE_LINES:
                flag = True
                continue
            if upper and NON_RETURNABLE_PHRASE in upper:
                flag = True
                continue
            filtered.append(entry)
        if flag:
            filtered = [entry for entry in filtered if entry.strip()]
        return filtered, flag

    def is_ocr_gibberish(line):
        """Enhanced gibberish detection for OCR artifacts like KKK RRR KEK patterns"""
        line_clean = line.strip()
        
        # Skip if too short
        if len(line_clean) < 5:
            return False
        
        # Pattern 1: Lines with mostly K, R, E combinations (your specific case)
        kre_pattern = re.sub(r'[^KRE\s]', '', line_clean.upper())
        if len(kre_pattern) > len(line_clean) * 0.7:  # 70% or more K/R/E chars
            return True
        
        # Pattern 2: Repetitive character patterns
        letters_only = re.sub(r'[^A-Z]', '', line_clean.upper())
        if len(letters_only) >= 10:
            # Count unique characters
            unique_chars = len(set(letters_only))
            # If 80% of the line is made up of only 2-3 characters, it's likely gibberish
            if unique_chars <= 3 and len(letters_only) > 15:
                return True
        
        # Pattern 3: Lines with very few vowels (original logic)
        vowels = sum(1 for c in letters_only if c in "AEIOU")
        return len(letters_only) >= 10 and vowels < 2
    
    def clean_gibberish_suffix(desc):
        """Remove OCR gibberish from the end of descriptions"""
        # Pattern 1: Remove trailing gibberish like "- KEKE KKK KEK KKK..."
        desc = re.sub(r'\s*-\s*[KRE\s]{10,}$', '', desc, flags=re.IGNORECASE)
        
        # Pattern 2: Remove any trailing section that's mostly K/R/E
        parts = desc.split(' - ')
        cleaned_parts = []
        
        for part in parts:
            kre_content = re.sub(r'[^KRE\s]', '', part.upper())
            # If this part is 80% or more K/R/E characters, skip it
            if len(kre_content) < len(part) * 0.8:
                cleaned_parts.append(part)
            else:
                # This part is gibberish, stop adding parts
                break
        
        return ' - '.join(cleaned_parts).strip()


    def clean_number(num_str):
        num_str = num_str.strip()
        if not num_str:
            return None
        parts = re.split(r"[.,]", num_str)
        if len(parts) == 1:
            return float(parts[0])
        decimal_part = parts[-1]
        integer_part = ''.join(parts[:-1])
        return float(f"{integer_part}.{decimal_part}")

    def is_junk_line(line):
        line_upper = line.upper()
        junk_indicators = [
            "CONTINUED", "MERCHANDISE", "SUBTOTAL", "TOTAL", "PAGE", "QUOTE", "DATE", "CUSTOMER", "SHIP",
            "TERMS", "TAX", "FREIGHT"
        ]
        end_of_doc_indicators = [
            "HARRINGTON", "TERMS AND CONDITIONS", "HIPCO.COM", "VOLATILITY", "TARIFF", "QUOTATION",
            "PURCHASE ORDER", "PROVISIONAL", "SUBJECT TO PRICE", "CURRENTLY", "VALID FOR 24 HOURS", "CONFIRMATION", "APPROVAL TO PROCEED"
        ]
        for indicator in junk_indicators + end_of_doc_indicators:
            if indicator in line_upper:
                return True
        non_alpha = sum(1 for c in line if not c.isalpha())
        alpha = sum(1 for c in line if c.isalpha())
        return (len(line) > 5 and non_alpha > alpha * 2) or len(line.strip()) < 3

    def normalize_line_ocr(line):
        replacements = {
            r"\bi\)": "9", r"\bI\)": "9", r"\bl\)": "9", r"\b\)\b": "9",
        }
        for pattern, replacement in replacements.items():
            line = re.sub(pattern, replacement, line)
        line = line.replace(")", "9").replace("(", "9")
        line = re.sub(r'[Oo](?=\d)', '0', line)
        line = re.sub(r'[Ll|](?=\d)', '1', line)
        return line

    def clean_line(text):
        return re.sub(r"[^\x00-\x7F]+", "", text).strip()

    def extract_info(lines_page1, lines_page2):
        company = ""
        quote_number = ""
        quote_date = ""
        expire_date = ""

        # Page 1: Get company name above email
        for idx, line in enumerate(lines_page1):
            line_clean = clean_line(line)
            if "@" in line_clean and "." in line_clean:
                for j in range(idx - 1, -1, -1):
                    prev_line = clean_line(lines_page1[j])
                    if prev_line:
                        company = re.sub(r"^\s*Company:\s*", "", prev_line, flags=re.IGNORECASE)
                        break
                break

        # Page 2: Get quote fields
        for line in lines_page2:
            line_clean = clean_line(line)
            if "Quote Date" in line_clean and not quote_date:
                match = re.search(r"Quote Date\s*([0-9/]+)", line_clean)
                if match:
                    quote_date = match.group(1)
            if "Expire Date" in line_clean and not expire_date:
                match = re.search(r"Expire Date\s*([0-9/]+)", line_clean)
                if match:
                    expire_date = match.group(1)
            if "Quotation#" in line_clean and not quote_number:
                match = QUOTE_NUMBER_PATTERN.search(line_clean)
                if match:
                    quote_number = match.group(0).upper()
                else:
                    fallback = re.search(r"Quotation#\s*([A-Z0-9\-]+)", line_clean, re.IGNORECASE)
                    if fallback:
                        quote_number = _normalize_quote_number(fallback.group(1))

        return company, quote_number, quote_date, expire_date

    def ocr_pages(pdf_file):
        bundle = ocr_bundle or get_ocr_bundle(pdf_file)
        return bundle.all_text_with_page, bundle.page_line_map

    # === OCR pages ===
    all_text_with_page, page_line_map = ocr_pages(PDF_FILE)

    # === Extract header info ===
    page1_lines = page_line_map.get(1, [])
    page2_lines = page_line_map.get(2, [])
    company, quote_num, quote_date, exp_date = extract_info(page1_lines, page2_lines)
    if not quote_num:
        for raw_line, _ in all_text_with_page:
            normalized = normalize_line_ocr(raw_line)
            match = QUOTE_NUMBER_PATTERN.search(normalized)
            if match:
                quote_num = _normalize_quote_number(match.group(0))
                break

    # === Set dynamic output file name ===
  

    customer_name = re.sub(r'[^\w\- ]', '_', (company or "UNKNOWN COMPANY").strip())
    safe_quote    = re.sub(r'[^\w\-]', '_', (quote_num or 'UNKNOWN'))
    OUTPUT_DIR    = ensure_dir(Path(OUTPUT_ROOT) / customer_name / "Quotes")
    OUTPUT_FILE   = OUTPUT_DIR / f"{safe_quote}_Quote.xlsx"

    # --- Updated Regex Pattern ---
    product_pattern = re.compile(
        r"([A-Za-z0-9\"/\.\-\s]+)\s+(\d+)\s+([\d,\.]+)\s+(EA|EBA|BA|FA|EFA|FT)\s+([\d,\.]+)"
    )

    # --- Parse lines ---
    non_returnable_flags = []
    i = 0
    while i < len(all_text_with_page):
        line, page_num = all_text_with_page[i]
        line = normalize_line_ocr(line)

        if not line or "continued" in line.lower() or "Product /Description Quantity Price U/M Extension" in line or "Merchandise" in line:
            i += 1
            continue

        # Skip lines that are just asterisks
        if re.match(r'^\s*\*+\s*$', line):
            i += 1
            continue

        # Starred lines
        if line.startswith("*"):
            line_clean = re.sub(r"^\*\d+\s*", "", line).strip()
            line_clean = normalize_line_ocr(line_clean)
            tokens = line_clean.split()

            desc_tokens = []
            numeric_start_idx = None
            for idx, tok in enumerate(tokens):
                if tok.strip().isdigit():
                    numeric_start_idx = idx
                    break
                desc_tokens.append(tok)
            first_line_desc = " ".join(desc_tokens)

            quantity = price = amount = None
            um_val = None
            if numeric_start_idx is not None and numeric_start_idx + 3 < len(tokens):
                qty_tok = re.sub(r"[^\d]", "", tokens[numeric_start_idx])
                quantity = int(qty_tok) if qty_tok else 0
                price = clean_number(tokens[numeric_start_idx + 1])
                um_val = tokens[numeric_start_idx + 2].upper()
                if um_val in ["EBA", "BA", "FA", "EFA"]:
                    um_val = "EA"
                amount = clean_number(tokens[numeric_start_idx + 3])

            i += 1
            if i >= len(all_text_with_page):
                break

            product_line, _ = all_text_with_page[i]
            product_line = normalize_line_ocr(product_line)
            tokens = product_line.split()
            product_id = tokens[0] if tokens else None
            # Clean up product ID by removing extra spaces
            if product_id:
                product_id = re.sub(r'(?<=\d)\s*\.\s*(?=\d)', '.', product_id.strip())
            else:
                product_id = ""
            remaining_desc = " ".join(tokens[1:]).strip() if len(tokens) > 1 else ""
            desc_lines = [first_line_desc]
            if remaining_desc:
                desc_lines.append(remaining_desc)

            j = i + 1
            while j < len(all_text_with_page):
                next_line, _ = all_text_with_page[j]
                next_line = normalize_line_ocr(next_line).replace("ASAHT", "ASAHI")
                if not next_line or is_ocr_gibberish(next_line):
                    j += 1
                    continue
                if CONTINUE_PAT.search(next_line):
                    break
                if product_pattern.search(next_line) or is_junk_line(next_line):
                    break
                desc_lines.append(next_line)
                j += 1

            desc_lines, non_returnable_flag = strip_non_returnable(desc_lines)
            primary_segment = desc_lines[0].strip() if desc_lines else ""
            full_description = " - ".join([segment for segment in desc_lines if segment]).strip(" -")
            all_products.append({
                "Page": page_num,
                "Product ID": product_id,
                "Description": full_description,
                "Quantity": quantity,
                "Price": price,
                "U/M": um_val,
                "Amount": amount,
                "Star Header": primary_segment
            })
            non_returnable_flags.append(non_returnable_flag)
            i = j
            continue

        # Standard product line
        match = product_pattern.search(line)
        if match:
            product_id_raw = match.group(1)
            # Clean up the product ID by removing trailing spaces but keeping internal spaces for dimensions
            product_id = re.sub(r'(?<=\d)\s*\.\s*(?=\d)', '.', product_id_raw.strip())

            quantity_raw = match.group(2).replace(")", "9").replace("(", "9")  # Fix OCR errors
            quantity = int(re.sub(r"[^\d]", "", quantity_raw) or "0")
            price = clean_number(match.group(3))
            um_val = match.group(4).upper()
            if um_val in ["EBA", "BA", "FA", "EFA"]:
                um_val = "EA"
            amount = clean_number(match.group(5))

            desc_lines = []
            j = i + 1
            while j < len(all_text_with_page):
                next_line, _ = all_text_with_page[j]
                next_line = normalize_line_ocr(next_line).replace("ASAHT", "ASAHI")
                if not next_line or is_ocr_gibberish(next_line):
                    j += 1
                    continue
                if CONTINUE_PAT.search(next_line):
                    break
                if product_pattern.search(next_line) or is_junk_line(next_line):
                    break
                desc_lines.append(next_line)
                j += 1

            desc_lines, non_returnable_flag = strip_non_returnable(desc_lines)
            description = " - ".join([segment for segment in desc_lines if segment]).strip(" -")
            all_products.append({
                "Page": page_num,
                "Product ID": product_id,
                "Description": description,
                "Quantity": quantity,
                "Price": price,
                "U/M": um_val,
                "Amount": amount,
                "Star Header": None
            })
            non_returnable_flags.append(non_returnable_flag)
            i = j
            continue

        i += 1

    # ===== COMPREHENSIVE CLEANUP BLOCK =====
    for product in all_products:
        desc = product.get("Description", "") or ""
        pid = product.get("Product ID", "") or ""



        while JUNK_SUFFIX_STRIPPER.search(desc):
            desc = JUNK_SUFFIX_STRIPPER.sub('', desc).strip(' -')

        # === ACID WASTE PIPE ID CORRECTION ===
        acid_patterns = [
            r'(\d{1,2})["\'\u201d\u2019]\s+PIPE\s+PP\s+FR\s+SCH40\s+.*ACID\s+WASTE',
            r'(\d{1,2})["\'\u201d\u2019]\s+.*PIPE\s+.*ACID\s+WASTE',
            r'(\d{1,2})\s*INCH\s+.*PIPE\s+.*ACID\s+WASTE'
        ]
        
        acid_match = None
        for pattern in acid_patterns:
            acid_match = re.search(pattern, desc, re.IGNORECASE)
            if acid_match:
                break
        
        if acid_match:
            size = int(acid_match.group(1))
            if 1 <= size <= 99:  # Support sizes 1-99
                product["Product ID"] = f"W{size:03d}"
                print(f"Acid waste pipe detected: {size}\" -> Product ID: W{size:03d}")

        # ===== BRAND NAME FIXES =====
        # 1. Fix common brand name OCR errors
        desc = desc.replace("ASAHT", "ASAHI")

        # ===== STANDARD/SPECIFICATION FIXES =====
        # 2. ANTSI ? ANSI (OCR variants)
        desc = re.sub(r'ANTSI', 'ANSI', desc, flags=re.IGNORECASE)

        # 3. ANSTI ? ANSI (OCR variants like ANST1, ANSTl)
        desc = re.sub(r'ANST[1iIl|]', 'ANSI', desc, flags=re.IGNORECASE)

        # ===== PRESSURE RATING FIXES =====
        # 4. Fix PSTI variants attached to numbers (150PSTI ? 150 PSI)
        desc = re.sub(r'(\d)PST[1iIl|]', r'\1 PSI', desc, flags=re.IGNORECASE)

        # 5. Fix standalone PSTI variants (PSTI ? PSI)
        desc = re.sub(r'\bPST[1iIl|]\b', 'PSI', desc, flags=re.IGNORECASE)

        # 6. Normalize PSI spacing (150PSI ? 150 PSI)
        desc = re.sub(r'(\d)PSI\b', r'\1 PSI', desc, flags=re.IGNORECASE)

        # ===== SDR (STANDARD DIMENSION RATIO) FIXES =====
        # 7. Fix SDR variants without spaces (SDR1il1, SDRlll ? SDR11)
        desc = re.sub(r'\bSDR[1ilI|]{1,6}\b', 'SDR11', desc, flags=re.IGNORECASE)

        # 8. Fix SDR variants with spaces (SDR 11, SDR 1l ? SDR11)
        desc = re.sub(r'\bSDR\s+[1ilI|]{1,6}\b', 'SDR11', desc, flags=re.IGNORECASE)

        # ===== SCHEDULE PIPE FIXES =====
        # 9. Fix SCH80 with OCR garbage (SCH80O, SCH80L ? SCH80)
        desc = re.sub(r'\bSCH80[^0-9\s]{1,3}\b', 'SCH80', desc, flags=re.IGNORECASE)

        # 10. Fix SCH40 with OCR garbage (SCH40O, SCH40L ? SCH40)
        desc = re.sub(r'\bSCH40[^0-9\s]{1,3}\b', 'SCH40', desc, flags=re.IGNORECASE)

        # ===== PN6 FLANGE RATING FIXES =====
        # 11. Fix parenthetical PN6 variants ((pn6), (png), ( pn6 ) ? PN6)
        desc = PN_BAD.sub(' PN6', desc)

        # 12. Add space before PN6 if missing (W/PN6 ? W/ PN6)
        desc = PN_JOIN.sub(r' \1', desc)
        desc = PN_TRAIL9.sub('PN6', desc)
        desc = PN_LEADING9.sub(r'\1', desc)

        # ===== PARENTHESIS CLEANUP =====
        # 13. Remove stray opening parentheses
        desc = desc.replace('(', '')

        # ===== PRO-90 ELBOW FIXES =====
        # 14. Fix common PRO-90 OCR variants
        desc = PRO_BAD_PATTERN.sub('PRO-90', desc)
        desc = PRO_SINGLE.sub('PRO-90', desc)
        desc = PRO_L_MIX.sub('PRO150', desc)
        desc = PRO_90_VARIANT.sub('PRO-90', desc)

        # 15. Normalize PRO-90 spacing (PRO -90, PRO- 90  PRO-90)
        desc = PRO_SINGLE.sub('PRO-90', desc)

        # --- Extra PRO-90 cleanup ---
        desc = re.sub(r'\bPR[O0Q]9[O0](?:[0O])?\b', 'PRO-90', desc, flags=re.IGNORECASE)
        desc = re.sub(r'\bPRO9[0O]{2}\b', 'PRO-90', desc, flags=re.IGNORECASE)

        # --- Remove stray 9s near PN6 ---
        desc = re.sub(r'\b9\s*(PN6)\b', r'\1', desc, flags=re.IGNORECASE)
        desc = re.sub(r'\b(PN6)\s*9\b', r'\1', desc, flags=re.IGNORECASE)

        # ===== PRO CODE DIMENSION FIXES =====
        # 16. Fix PRO codes with OCR errors in numbers (PRO1L50xPRO150 ? PRO150xPRO150)
        desc = PRO_PAIR.sub(lambda m: f"PRO{PRO_NUM_FIX.sub('1', m.group(1))}x{m.group(2)}", desc)

        # 17. Fix malformed PRO connections (PRO123XABC456 ? PRO123x456)
        desc = PRO_DOUBLE.sub(r'\1x\2', desc)

        # ===== DIMENSION FORMATTING FIXES =====
        # 18. Normalize inch formatting (4"x%1/2" ? 4"x1/2")
        desc = DIM_RATIO.sub(r'\1x\2', desc)

        # 19. Collapse repeated X's (xx, xX, Xx ? x)
        desc = DIM_MULTI_X.sub(r'\1', desc)

        # 20. Clean up junk between dimensions (PRO150xX150 ? PRO150x150)
        desc = DIM_X_BETWEEN.sub(r'\1x', desc)

        # 21. Remove dash after trailing x (PRO150x - PRO90 ? PRO150xPRO90)
        desc = DIM_SPACE_DASH.sub(r'x\1', desc)

        desc = re.sub(r'(?<=\d)\s*-\s*(?=\d)', ' - ', desc)

        desc = desc.replace("!", "'")

        




        pid_clean = pid
        if PRO_L_MIX.search(pid_clean):
            pid_clean = PRO_L_MIX.sub('PRO150', pid_clean)
        if PRO_90_VARIANT.search(pid_clean):
            pid_clean = PRO_90_VARIANT.sub('PRO90', pid_clean)

        desc_upper = desc.upper()
        if 'FLANGE BLIND' in desc_upper and pid_clean.endswith('8'):
            pid_clean = pid_clean[:-1] + 'S'
        if (' S ' in desc_upper or PRO_SLIP.search(desc)) and not pid_clean.endswith('S'):
            pid_clean = PID_SUFFIX.sub('S', pid_clean)

        product["Product ID"] = pid if pid.startswith("W") else pid_clean

        if product.get("Star Header") is not None:
            primary_clean = desc.split(" - ")[0].strip() if desc else ""
            product["Star Header"] = primary_clean or product["Star Header"]

        # ===== SAVE CLEANED DESCRIPTION =====
        product["Description"] = desc



    # --- Export to Excel ---
    df = pd.DataFrame(all_products)
    star_headers = df["Star Header"].tolist() if "Star Header" in df.columns else [None] * len(df)
    if "Star Header" in df.columns:
        df = df.drop(columns=["Star Header"])
    df.to_excel(OUTPUT_FILE, index=False, startrow=7)

  #  from tkinter import messagebox
  #  messagebox.showinfo("Save Path", f"Saved to:\n{OUTPUT_FILE}")


    print(">>> File created at:", OUTPUT_FILE)
    print(">>> Customer folder:", Path(OUTPUT_FILE).parent)

    # === Format Excel ===
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Column widths
    for col in ['A', 'B', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['C'].width = 125

    # Header style
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[8]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = border

    # Data formatting
    currency_format = '"$"#,##0.00'
    non_returnable_fill = PatternFill(start_color="FFECC6", end_color="FFECC6", fill_type="solid")
    for row_idx, row in enumerate(ws.iter_rows(min_row=9, max_row=ws.max_row, max_col=7)):
        for i, cell in enumerate(row):
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if i == 4 or i == 6:
                cell.number_format = currency_format
        if non_returnable_flags and row_idx < len(non_returnable_flags) and non_returnable_flags[row_idx]:
            for cell in row:
                cell.fill = non_returnable_fill

    # Bold primary description for starred products
    if star_headers:
        for idx, header in enumerate(star_headers):
            if not header or not isinstance(header, str):
                continue
            excel_row = 9 + idx
            if excel_row > ws.max_row:
                break
            cell = ws.cell(row=excel_row, column=3)
            full_text = cell.value or ""
            primary_text = header.strip()
            if not full_text or not primary_text:
                continue
            if full_text.startswith(primary_text):
                first_part = primary_text
                rest = full_text[len(primary_text):]
            elif full_text.lower().startswith(primary_text.lower()):
                first_part = full_text[:len(primary_text)]
                rest = full_text[len(primary_text):]
            elif " - " in full_text:
                first_part = full_text.split(" - ", 1)[0]
                rest = full_text[len(first_part):]
            else:
                continue
            rich_text = CellRichText()
            rich_text = CellRichText()
            rich_text.append(TextBlock(InlineFont(b=True), first_part))
            if rest:
                rich_text.append(TextBlock(InlineFont(), rest))
            cell.value = rich_text
    # Freeze header row
    ws.freeze_panes = 'A9'

    # Remove gridlines
    ws.sheet_view.showGridLines = False

    # === OCR Extract from Page 1 & 2 ===


    # === Fallbacks ===
    subject = quote_num or "UNKNOWN"
    company_name = company or "UNKNOWN COMPANY"
    quote_date = quote_date or "MM/DD/YYYY"
    expire_date = exp_date or "MM/DD/YYYY"

    # === Create customer folder ===
   # base_dir = Path(OUTPUT_ROOT)
   # customer_name = re.sub(r'[^\w\- ]', '_', company_name.strip())
  #  customer_folder = ensure_dir(Path(OUTPUT_ROOT) / customer_name / "Quotes")




    # === Set final output file path inside customer folder ===
 #   safe_quote = re.sub(r'[^\w\-]', '_', quote_num or 'UNKNOWN')
 #   OUTPUT_FILE = customer_folder / f"{safe_quote}_Quote.xlsx"

    for i in range(1, 8):
        ws.row_dimensions[i].height = 18 if i >= 3 else 35
    ws.row_dimensions[7].height = 6

    # === Column Widths ===
    widths = {"A": 15, "B": 15, "C": 106, "D": 15, "E": 15, "F": 15, "G": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # === Merged Cells ===
    ws.merge_cells("C1:C5")  # Logo
    ws.merge_cells("E3:G3")  # Company
    ws.merge_cells("E4:G4")  # Quote #
    ws.merge_cells("E5:G5")  # Quote Date
    ws.merge_cells("E6:G6")  # Exp Date

    # === Labels ===
    labels = ["Customer", "Quotation#", "Quote Date", "Exp Date"]
    for i, label in enumerate(labels, start=3):
        cell = ws[f"D{i}"]
        cell.value = label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # === Values ===
    values = [company_name, subject, quote_date, expire_date]
    for i, val in enumerate(values, start=3):
        ws[f"E{i}"].value = val
        for col in ["E", "F", "G"]:
            ws[f"{col}{i}"].alignment = Alignment(horizontal="left", vertical="center")

    # === Logo ===
    if LOGO_PATH and LOGO_PATH.exists():
        logo = XLImage(str(LOGO_PATH))
        logo.width *= 0.82
        logo.height *= 0.82
        logo.anchor = "C1"
        ws.add_image(logo)

    # === Style ===
    ws.sheet_view.showGridLines = False
    # === Add Total Line for Amount ===
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right', vertical='center')
    total_cell = ws.cell(row=total_row, column=7)
    last_data_row = ws.max_row - 1
    total_cell.value = f"=SUM(G9:G{last_data_row})"
    total_cell.number_format = '"$"#,##0.00'
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
    for col in range(1, 8):  # A (1) to H (8)
        ws.cell(row=total_row, column=col).border = thin_border

    # Save file
    wb.save(OUTPUT_FILE)
    
    print("Quote extracted and report saved.")
    return OUTPUT_FILE
