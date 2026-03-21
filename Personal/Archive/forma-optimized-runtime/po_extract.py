from pathlib import Path
from typing import Sequence
import re
import pytesseract
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from config import (
    TESSERACT_PATH,
    POPPLER_PATH,
    LOGO_IMAGE,
    POS_DIR,
    ensure_dir
)
from ocr_service import OCRBundle, get_ocr_bundle

PO_NUMBER_PATTERN = re.compile(r'\b\d{3}[A-Za-z]{2}\d{3}\b')
PO_NUMBER_LABEL_REGEX = re.compile(r'P[\\/]?O\s*Number[:\s]*([A-Za-z0-9-]+)', re.IGNORECASE)
PURCHASE_ORDER_FALLBACK_REGEX = re.compile(r"Purchase\s*Order[^A-Za-z0-9]+([A-Za-z0-9\-]+)", re.IGNORECASE)
CUSTOMER_PO_REGEX = re.compile(r"Cust\s*P\/O#\s*([^\s]+)", re.IGNORECASE)
PO_DATE_REGEX = re.compile(r"PO-?Date[:\s]*([0-9/]+)", re.IGNORECASE)
CONTACT_FREIGHT_REGEX = re.compile(r"Contact:(.*?)Freight Terms:(.*)", re.IGNORECASE)
PAYMENT_TERMS_REGEX = re.compile(r"Pymnt\s*Terms:\s*(.*)", re.IGNORECASE)
INSTRUCTIONS_START_REGEX = re.compile(r"^\s*Instructions:", re.IGNORECASE)
TABLE_HEADER_REGEX = re.compile(r"Qty-Opn.*Product\s*/\s*Description.*U/M.*Cost.*Req-?Date", re.IGNORECASE)
INSTRUCTIONS_STOP_REGEX = re.compile(r"(Qty|QTY).{0,30}(Product|Description|U/M|Cost)", re.IGNORECASE)
SHIP_TO_REGEX = re.compile(r"\bShip\s*To:?\b", re.IGNORECASE)

def sanitize_name(s):
    return re.sub(r'[^\w\- ]', '_', s.strip()) or "UNKNOWN"

def _normalize_so_noise(s: str) -> str:
    return re.sub(r'\bS[\s$/_-]*-\b', 'S-', s)



def _normalize_po_number(value: str) -> str:
    if not value:
        return ""

    candidate = value.upper()
    match = PO_NUMBER_PATTERN.search(candidate)
    if match:
        token = match.group(0)
    else:
        token = re.sub(r"[^A-Za-z0-9]", "", candidate)

    if len(token) >= 5:
        prefix = token[:3]
        letters = token[3:5].replace('0', 'O')
        suffix = token[5:]
        token = prefix + letters + suffix

    return token



def clean_description_and_product_id(product):
    # Compile regex inside the function for modularity
    PRO_SLIP = re.compile(r"\bSLIP\b", re.I)
    PRO_BAD_PATTERN = re.compile(r"PRO[-\s]?9[0O]", re.I)
    PRO_L_MIX = re.compile(r"PRO[1lI|]{1,2}50", re.I)
    ANSI1 = re.compile(r"ANTSI", re.I)
    ANSI2 = re.compile(r"ANST[1iIl|]", re.I)
    PSI_FIX = re.compile(r"(\d)PSI\b", re.I)
    SDR_FIX = re.compile(r"SDR\s*[1ilI|]{1,6}", re.I)
    SCH80 = re.compile(r"SCH80[^0-9\s]{1,3}", re.I)
    SCH40 = re.compile(r"SCH40[^0-9\s]{1,3}", re.I)
    PN_BAD = re.compile(r"\(\s*P\s*N\s*[6GHB9QOD]\s*\)", re.I)
    PN_JOIN = re.compile(r"(?<!\s)(PN6)", re.I)
    PN_TRAIL9 = re.compile(r"\bPN([6GHB9QOD])9\b", re.I)
    PN_LEAD9 = re.compile(r"\b9\s*(PN6)\b", re.I)

    desc = (product.get("Description") or "").strip()
    pid = re.sub(r"(?<=\d)\s+(?=\d)", "", (product.get("Product ID") or "").strip())

    if not desc:
        return product

    desc = desc.replace("ASAHT", "ASAHI")
    desc = ANSI1.sub("ANSI", desc)
    desc = ANSI2.sub("ANSI", desc)
    desc = PSI_FIX.sub(r"\1 PSI", desc)
    desc = SDR_FIX.sub("SDR11", desc)
    desc = SCH80.sub("SCH80", desc)
    desc = SCH40.sub("SCH40", desc)
    desc = PN_BAD.sub(" PN6", desc)
    desc = PN_JOIN.sub(r" \1", desc)
    desc = PN_TRAIL9.sub("PN6", desc)
    desc = PN_LEAD9.sub("PN6", desc)
    desc = PRO_BAD_PATTERN.sub("PRO-90", desc)
    desc = PRO_L_MIX.sub("PRO150", desc)

    length_matches = list(re.finditer(r"\bLENGTH\b", desc, re.IGNORECASE))
    if length_matches and length_matches[-1].end() < len(desc):
        desc = desc[:length_matches[-1].end()].rstrip(" -/,")

    if PRO_L_MIX.search(pid):
        pid = PRO_L_MIX.sub("PRO150", pid)

    if "FLANGE BLIND" in desc.upper() and pid.endswith("8"):
        pid = pid[:-1] + "S"
    if " S " in desc.upper() and not pid.endswith("S"):
        pid = re.sub(r"([A-Za-z0-9])$", "S", pid)

    pipe_match = re.search(r"\b(\d{1,2})(?:\"|\')?\s+PIPE\s+PP\s+FR\s+SCH40\s+.*ACID\s+WASTE", desc, re.I)
    if pipe_match:
        size = int(pipe_match.group(1))
        if 2 <= size <= 20:
            pid = f"W{size:03}"

    product["Product ID"] = pid
    product["Description"] = desc
    return product

def po_extract(file_path, output_root=None, ocr_bundle: OCRBundle | None = None):
    pytesseract.pytesseract.tesseract_cmd = str(TESSERACT_PATH)
    PDF_FILE = Path(file_path)
    OUTPUT_ROOT = ensure_dir(Path(output_root)) if output_root else ensure_dir(POS_DIR)

    if ocr_bundle is not None:
        all_lines = list(ocr_bundle.all_lines)
    else:
        bundle = get_ocr_bundle(PDF_FILE)
        all_lines = list(bundle.all_lines)

    def find_after(pattern):
        for line in all_lines:
            match = pattern.search(line) if hasattr(pattern, "search") else re.search(pattern, line, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        return ""

    def next_nonempty_idx(start_idx):
        for i in range(start_idx + 1, len(all_lines)):
            if all_lines[i].strip():
                return i
        return None

    def split_vendor_customer(line):
        suffixes = [" INC", " LLC", " LLP", " LP", " CORP", " CORPORATION", " CO", " COMPANY", " LTD", " LIMITED"]
        for suffix in suffixes:
            idx = line.upper().find(suffix)
            if idx != -1 and idx + len(suffix) < len(line):
                return line[:idx + len(suffix)].strip(), line[idx + len(suffix):].strip()
        parts = re.split(r"\s{2,}", line)
        return parts[0], parts[1] if len(parts) > 1 else ""

    po_number = ""
    customer_po = ""
    po_date = ""
    contact = ""
    freight_terms = ""
    payment_terms = ""
    instructions = ""
    vendor_name = ""
    customer_name = ""
    instruction_start_idx = None
    table_start_idx = None
    ship_to_idx = None

    for idx, line in enumerate(all_lines):
        if not po_number:
            match = PO_NUMBER_LABEL_REGEX.search(line) or PURCHASE_ORDER_FALLBACK_REGEX.search(line)
            if match:
                po_number = match.group(1).strip()
        if not customer_po:
            match = CUSTOMER_PO_REGEX.search(line)
            if match:
                customer_po = match.group(1).strip()
        if not po_date:
            match = PO_DATE_REGEX.search(line)
            if match:
                po_date = match.group(1).strip()
        if not contact and "Contact:" in line and "Freight Terms:" in line:
            match = CONTACT_FREIGHT_REGEX.search(line)
            if match:
                contact = match.group(1).strip()
                freight_terms = match.group(2).strip()
        if not payment_terms:
            match = PAYMENT_TERMS_REGEX.search(line)
            if match:
                payment_terms = re.sub(r"\bFOB:.*", "", match.group(1)).strip()
        if instruction_start_idx is None and INSTRUCTIONS_START_REGEX.match(line):
            instruction_start_idx = idx
        if table_start_idx is None and TABLE_HEADER_REGEX.search(line):
            table_start_idx = idx + 1
        if ship_to_idx is None and SHIP_TO_REGEX.search(line):
            ship_to_idx = idx

    po_number = _normalize_po_number(po_number)

    if instruction_start_idx is not None:
        collected = [re.sub(r"^\s*Instructions:\s*", "", all_lines[instruction_start_idx], flags=re.IGNORECASE).strip()]
        for j in range(instruction_start_idx + 1, len(all_lines)):
            if INSTRUCTIONS_STOP_REGEX.search(all_lines[j]):
                break
            collected.append(all_lines[j])
        instructions = " | ".join(part for part in collected if part)

    if ship_to_idx is not None:
        next_idx = next_nonempty_idx(ship_to_idx)
        if next_idx is not None:
            vendor_name, customer_name = split_vendor_customer(all_lines[next_idx])

    # === Product Table ===
    products = []
    line_pat = re.compile(
        r"^(?P<qty>\d+)\s+(?P<pid>[A-Za-z0-9\-]+)\s+(?P<uom>[A-Z]{2,4})\s+(?P<price>[\d.,]+)\s+(?P<req>\d{2}/\d{2}/\d{2})$"
    )
    sales_order_pat = re.compile(r"\bS-[A-Za-z0-9]{8}\b")
    amount_pat = re.compile(r"\b([\d,]+\.\d{2})\b")
    def _is_table_stop(line: str) -> bool:
        lower = line.lower()
        return (
            lower.startswith("subtotal")
            or lower.startswith("total")
            or lower.startswith("merchandise")
            or lower.startswith("shipping")
            or lower.startswith("tax")
            or lower.startswith("amount due")
            or "terms and conditions" in lower
            or "harrington's standard" in lower
        )

    header_prefixes = (
        "** purchase order",
        "page ",
        "fremont branch",
        "harrington industrial",
        "this purchase order supersedes",
        "po-date",
        "contact:",
        "pymnt terms",
        "instructions:",
        "qty-opn",
    )

    header_contains = (
        "purchase order",
        "regular p/o",
        "ship to",
        "harrington location",
        "buyer:",
        "ship via",
        "freight terms",
        "fob",
        "po box",
        "city of industry",
        "fremont, ca",
        "fremont",
        "branch",
        "page",
        "arrington",
        "arringten",
        "davenport",
        "supersedes",
        "cust p/o",
        "customer p/o",
        "customer po",
        "p o box",
        "acco air conditioning",
        "job#",
        "cust p/o#",
    )

    def _is_page_header(line: str) -> bool:
        lower = line.lower()
        return any(lower.startswith(prefix) for prefix in header_prefixes) or any(token in lower for token in header_contains)

    start_idx = table_start_idx

    if start_idx is not None:
        i = start_idx
        while i < len(all_lines):
            line = all_lines[i]
            if _is_table_stop(line):
                break
            match = line_pat.match(line)
            if not match:
                i += 1
                continue
            qty = match.group("qty")
            pid = match.group("pid")
            uom = match.group("uom")
            price_text = match.group("price").replace(',', '')
            req_date = match.group("req")
            try:
                price_value = float(price_text)
            except ValueError:
                i += 1
                continue

            d1_idx = next_nonempty_idx(i)
            desc_lines = []
            amount = ""
            sales_orders = []

            if d1_idx is not None:
                first_desc = _normalize_so_noise(all_lines[d1_idx])
                if not (
                    _is_table_stop(first_desc)
                    or re.search(r"(?i)\bcontinued\b", first_desc)
                    or _is_page_header(first_desc)
                    or re.search(r"^V\.?\s*PN#?", first_desc, re.IGNORECASE)
                ):
                    amt_match = amount_pat.search(first_desc)
                    if amt_match:
                        amount = amt_match.group(1)
                        first_desc = re.sub(r"\b[\d,]+\.\d{2}\b", "", first_desc).strip()
                    if first_desc:
                        desc_lines.append(first_desc)

            j = (d1_idx + 1) if d1_idx is not None else (i + 1)
            while j < len(all_lines):
                next_line = _normalize_so_noise(all_lines[j].strip())
                if not next_line:
                    j += 1
                    continue
                if line_pat.match(next_line) or _is_table_stop(next_line):
                    break
                if _is_page_header(next_line):
                    break
                if re.search(r"(?i)\bcontinued\b", next_line):
                    j += 1
                    break
                if re.search(r"^V\.?\s*PN#?", next_line, re.IGNORECASE):
                    j += 1
                    continue

                so_matches = sales_order_pat.findall(next_line)
                if so_matches:
                    for so in so_matches:
                        if so not in sales_orders:
                            sales_orders.append(so)
                        next_line = next_line.replace(so, "").strip()

                amt_match = amount_pat.search(next_line)
                if amt_match and not amount:
                    amount = amt_match.group(1)
                    next_line = next_line.replace(amount, "").strip()

                if next_line:
                    desc_lines.append(next_line)

                j += 1

            if not amount:
                try:
                    amount = f"{float(qty) * price_value:.2f}"
                except ValueError:
                    amount = ""

            description = " ".join(desc_lines).strip()
            description = re.sub(r"\s+", " ", description)
            sales_order_text = " | ".join(sales_orders)

            product = {
                "Quantity": qty,
                "Product ID": pid,
                "Description": description,
                "Sales Order": sales_order_text,
                "U/M": uom,
                "Price": price_value,
                "Amount": float(amount.replace(',', '')) if amount else "",
                "Req Date": req_date
            }

            product = clean_description_and_product_id(product)
            products.append(product)
            i = j

    total_amount = 0.0
    for product in products:
        value = product.get("Amount")
        if isinstance(value, (int, float)):
            total_amount += value
    total_row = {
        "Quantity": "",
        "Product ID": "",
        "Description": "TOTAL",
        "Sales Order": "",
        "U/M": "",
        "Price": "",
        "Amount": total_amount,
        "Req Date": ""
    }
    products.append(total_row)

    # === Export to Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    headers = [
        "Customer", "Purchase Order #", "Customer PO #", "PO Date",
        "Contact", "Freight Terms", "Payment Terms", "Instructions", "Vendor"
    ]
    values = [
        customer_name, po_number, customer_po, po_date,
        contact, freight_terms, payment_terms, instructions, vendor_name
    ]

    for i, (label, val) in enumerate(zip(headers, values), start=3):
        ws[f"D{i}"] = label
        ws[f"D{i}"].font = Font(bold=True)
        ws[f"D{i}"].alignment = Alignment(horizontal="right")
        ws[f"E{i}"] = val
        ws[f"E{i}"].alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=7)

    if LOGO_IMAGE.exists():
        logo = XLImage(str(LOGO_IMAGE))
        logo.width *= 0.82
        logo.height *= 0.82
        logo.anchor = "C1"
        ws.add_image(logo)

    table_headers = ["Quantity", "Product ID", "Description", "Sales Order", "U/M", "Price", "Amount", "Req Date"]
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, name in enumerate(table_headers, start=1):
        cell = ws.cell(row=12, column=col_idx, value=name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, product in enumerate(products, start=13):
        row_data = [product[h] for h in table_headers]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if col_idx in [6, 7] and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0.00'

    total_row_idx = ws.max_row
    if total_row_idx >= 13:
        for col_idx in range(1, len(table_headers) + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.font = Font(bold=True)
            if col_idx == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.freeze_panes = "A13"
    ws.sheet_view.showGridLines = False

    for i, w in zip(range(1, 9), [10, 18, 106, 20, 12, 15, 15, 15]):
        ws.column_dimensions[chr(64 + i)].width = w

    # --- Final path & save ---
    customer_name_clean = sanitize_name(customer_name or "UNKNOWN")
    po_number_clean     = sanitize_name(po_number or "UNKNOWN")

    # Build the full target path FIRST
    final_path = Path(OUTPUT_ROOT) / customer_name_clean / "POs" / f"{po_number_clean}_PO.xlsx"

    # Make sure the folder exists (creates .../[Customer]/POs)
    ensure_dir(final_path.parent)

    # Save
    wb.save(final_path)
    print("PO extracted and saved to: " + str(final_path))
    return final_path
