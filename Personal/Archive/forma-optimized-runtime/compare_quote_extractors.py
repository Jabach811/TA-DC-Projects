from __future__ import annotations

import importlib.util
import sys
import time
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
FORMA_ROOT = Path(r"C:\Users\mabac\OneDrive\Desktop\Joel\Hipco Project\Forma -  Main Folder")
DEFAULT_SAMPLE_PDF = FORMA_ROOT / "172451.pdf"
OUTPUT_ROOT = ROOT / "test-output"


def load_module(module_name: str, path: Path):
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def read_sheet_values(path: Path) -> list[list[object]]:
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    values = []
    for row in ws.iter_rows(values_only=True):
        values.append(list(row))
    return values


def normalize_data_rows(path: Path) -> list[list[object]]:
    rows = read_sheet_values(path)
    trimmed = []
    for row in rows[7:]:
        materialized = list(row[:7])
        if not any(value is not None and value != "" for value in materialized):
            continue
        descriptor = materialized[2]
        if isinstance(descriptor, str) and descriptor.strip().upper() == "TOTAL":
            continue
        trimmed.append(materialized)
    return trimmed


def resolve_output_path(run_result, output_root: Path) -> Path:
    if run_result:
        return Path(run_result)
    candidates = sorted(output_root.rglob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"No workbook found under {output_root}")
    return candidates[0]


def compare_workbooks(left: Path, right: Path) -> tuple[bool, str]:
    left_values = normalize_data_rows(left)
    right_values = normalize_data_rows(right)
    if left_values == right_values:
        return True, "Workbook cell values match exactly."

    max_rows = max(len(left_values), len(right_values))
    for row_idx in range(max_rows):
        left_row = left_values[row_idx] if row_idx < len(left_values) else []
        right_row = right_values[row_idx] if row_idx < len(right_values) else []
        if left_row != right_row:
            return (
                False,
                f"First mismatch at row {row_idx + 1}: original={left_row} optimized={right_row}",
            )
    return False, "Workbook comparison failed for an unknown reason."


def main() -> None:
    sys.path.insert(0, str(FORMA_ROOT))
    sample_pdf = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SAMPLE_PDF
    original = load_module("forma_original_quote_extractor", FORMA_ROOT / "quote_extractor.py")
    optimized = load_module("forma_optimized_quote_extractor", ROOT / "quote_extractor.py")

    original_output_root = OUTPUT_ROOT / "original"
    optimized_output_root = OUTPUT_ROOT / "optimized"
    original_output_root.mkdir(parents=True, exist_ok=True)
    optimized_output_root.mkdir(parents=True, exist_ok=True)

    start = time.perf_counter()
    original_error = None
    try:
        original_result = original.quotation_extract(str(sample_pdf), output_root=str(original_output_root))
    except Exception as exc:  # noqa: BLE001
        original_result = None
        original_error = exc
    original_path = resolve_output_path(original_result, original_output_root)
    original_seconds = time.perf_counter() - start

    start = time.perf_counter()
    optimized_result = optimized.quotation_extract(str(sample_pdf), output_root=str(optimized_output_root))
    optimized_path = resolve_output_path(optimized_result, optimized_output_root)
    optimized_seconds = time.perf_counter() - start

    matches, message = compare_workbooks(original_path, optimized_path)

    print(f"Original output:  {original_path}")
    print(f"Optimized output: {optimized_path}")
    print(f"Original time:    {original_seconds:.2f}s")
    print(f"Optimized time:   {optimized_seconds:.2f}s")
    print(f"Comparison:       {message}")
    if original_error is not None:
        print(f"Original status:  crashed with {type(original_error).__name__}: {original_error}")
    if original_seconds:
        improvement = ((original_seconds - optimized_seconds) / original_seconds) * 100
        print(f"Runtime delta:    {improvement:.1f}%")
    print(f"Matched values:   {matches}")


if __name__ == "__main__":
    main()
